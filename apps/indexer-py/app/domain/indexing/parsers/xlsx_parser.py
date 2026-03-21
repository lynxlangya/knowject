from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from app.domain.indexing.segments import ParsedSegment


ErrorFactory = Callable[[str], Exception]


def parse_xlsx_segments(
    *,
    path: Path,
    error_factory: ErrorFactory,
    clean_text: Callable[[str], str],
) -> list[ParsedSegment]:
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as error:  # pragma: no cover - library specific errors
        raise error_factory(f"XLSX 解析失败：{error}") from error

    segments: list[ParsedSegment] = []

    for sheet in workbook.worksheets:
        rows: list[tuple[int, tuple[object, ...], tuple[str, ...]]] = []

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            raw_values = tuple(row)
            values = tuple(_cell_to_text(cell) for cell in raw_values)
            if not any(values):
                continue
            rows.append((row_index, raw_values, values))

        if not rows:
            continue

        first_row_index, first_raw_values, first_values = rows[0]
        use_first_row_as_header = _should_use_first_row_as_header(
            first_raw_values,
            next_row=rows[1][1] if len(rows) > 1 else None,
        )
        header_values = (
            _normalize_header_values(first_values)
            if use_first_row_as_header
            else _build_default_headers(len(first_values))
        )
        data_rows = rows[1:] if use_first_row_as_header else rows

        for row_index, _raw_values, values in data_rows:
            if row_index == first_row_index and use_first_row_as_header:
                continue

            line_items: list[str] = []
            for index, value in enumerate(values):
                if not value:
                    continue
                header = (
                    header_values[index]
                    if index < len(header_values) and header_values[index]
                    else f"Column {index + 1}"
                )
                line_items.append(f"{header}: {value}")

            if not line_items:
                continue

            text = clean_text(f"{sheet.title} | 第 {row_index} 行\n" + "\n".join(line_items))
            if not text:
                continue

            segments.append(
                ParsedSegment(
                    text=text,
                    source_kind="xlsx",
                    order=len(segments),
                    sheet_name=sheet.title,
                    row_index=row_index,
                    header_path=header_values,
                )
            )

    workbook.close()
    return segments


def _should_use_first_row_as_header(
    first_row: tuple[object, ...],
    *,
    next_row: tuple[object, ...] | None,
) -> bool:
    if next_row is None:
        return False

    non_empty_values = [value for value in first_row if value is not None and str(value).strip()]
    if not non_empty_values:
        return False

    return all(isinstance(value, str) for value in non_empty_values)


def _normalize_header_values(values: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(
        value if value else f"Column {index + 1}" for index, value in enumerate(values)
    )


def _build_default_headers(column_count: int) -> tuple[str, ...]:
    return tuple(f"Column {index + 1}" for index in range(column_count))


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
